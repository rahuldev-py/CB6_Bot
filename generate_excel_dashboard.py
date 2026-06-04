"""
CB6 QUANTUM â€” Excel Dashboard Generator v3  (SaaS Dark Theme, fixed charts)
Reads  : data/cb6_master_archive.csv
Outputs: data/CB6_Dashboard.xlsx

Fixes vs v2:
  - chart.visible_cells_only = False  â†’  charts render even from narrow columns
  - plot area transparent via ShapeProperties(noFill)
  - chart data columns set to width 0.5 (never hidden â€” hidden blocks rendering)
  - non-overlapping charts: chart1 anchored B26, chart2 anchored K26
  - row25 labels aligned with their chart columns
"""

import os, csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import AreaChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE_DIR = os.path.dirname(__file__)
CSV_PATH = os.path.join(BASE_DIR, 'data', 'cb6_master_archive.csv')
OUT_PATH = os.path.join(BASE_DIR, 'data', 'CB6_Dashboard.xlsx')

# â”€â”€ Palette â€” all 6-char RGB (no alpha prefix, which corrupts chart XML) â”€â”€â”€â”€â”€â”€â”€
C_BG     = '0F172A'
C_BANNER = '060D1F'
C_CARD   = '1E293B'
C_HDR    = '0B1526'
C_GREEN  = '10B981'
C_RED    = 'F43F5E'
C_BLUE   = '3B82F6'
C_AMBER  = 'F59E0B'
C_WHITE  = 'F1F5F9'
C_MUTED  = '64748B'
C_BORDER = '374151'
C_ROW_D  = '0D1B2E'
C_ROW_L  = '1A2B42'
C_ORANGE = 'F97316'
C_PURPLE = 'A855F7'
C_CYAN   = '06B6D4'

INDEX_ACCENT = {
    'NIFTY':      C_BLUE,
    'BANKNIFTY':  C_ORANGE,
    'FINNIFTY':   C_PURPLE,
    'MIDCPNIFTY': C_CYAN,
}

# â”€â”€ Column layout â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# A(1)=margin | B(2)C(3)=card1 | D(4)=gap | E(5)F(6)=card2 | G(7)=gap |
# H(8)I(9)=card3 | J(10)=gap | K(11)L(12)=card4 | M(13)=gap |
# N(14)O(15)=card5 | P(16)=gap | Q(17)R(18)=card6 | S(19)=margin
CARD_PAIRS = [(2,3),(5,6),(8,9),(11,12),(14,15),(17,18)]
COL_WIDTHS = {
    1:1.5, 2:9.5, 3:9.5,
    4:1.2, 5:9.5, 6:9.5,
    7:1.2, 8:9.5, 9:9.5,
    10:1.2, 11:9.5, 12:9.5,
    13:1.2, 14:9.5, 15:9.5,
    16:1.2, 17:9.5, 18:9.5,
    19:1.5,
    # chart data helper cols â€” NARROW, never hidden (hidden cols break chart rendering)
    20:0.5, 21:0.5, 22:0.5, 23:0.5, 24:0.5, 25:0.5, 26:0.5,
}

# â”€â”€ Style helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _fill(c):
    return PatternFill('solid', fgColor=c)

def _font(size=10, bold=False, color=C_WHITE, italic=False, name='Calibri'):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def _align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _fill_range(ws, r1, c1, r2, c2, color):
    f = _fill(color)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).fill = f

def _outer_border(ws, r1, c1, r2, c2, color=C_BORDER, style='thin'):
    s    = Side(border_style=style, color=color)
    none = Side(border_style=None)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = Border(
                top    = s    if r == r1 else none,
                bottom = s    if r == r2 else none,
                left   = s    if c == c1 else none,
                right  = s    if c == c2 else none,
            )

def _mc(ws, r1, c1, r2, c2, value='', fill=None, font=None, align=None):
    cl = get_column_letter
    ws.merge_cells(f'{cl(c1)}{r1}:{cl(c2)}{r2}')
    cell = ws.cell(r1, c1)
    if value is not None: cell.value = value
    if fill  is not None: cell.fill  = fill
    if font  is not None: cell.font  = font
    if align is not None: cell.alignment = align
    return cell

def _setup_cols(ws):
    for ci, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

def _row_h(ws, heights):
    for r, h in heights.items():
        ws.row_dimensions[r].height = h

# â”€â”€ Chart transparency â€” openpyxl-native, no raw lxml required â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _transparent_plot_area(chart):
    """Set chart plot area to no-fill (transparent background)."""
    from openpyxl.chart.shapes import GraphicalProperties
    chart.plot_area.spPr = GraphicalProperties(noFill=True)

# â”€â”€ KPI card helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _kpi_card(ws, c1, c2, r, label, value, sublabel, accent, val_color):
    """5-row glassmorphism card: accent stripe | label | VALUE | sub | pad"""
    _fill_range(ws, r,   c1, r,   c2, accent)
    _fill_range(ws, r+1, c1, r+4, c2, C_CARD)
    _mc(ws, r+1, c1, r+1, c2, label,
        font=_font(7, False, C_MUTED), align=_align())
    _mc(ws, r+2, c1, r+2, c2, value,
        font=_font(20, True, val_color), align=_align())
    _mc(ws, r+3, c1, r+3, c2, sublabel,
        font=_font(7, False, C_MUTED, italic=True), align=_align())
    _mc(ws, r+4, c1, r+4, c2, '')
    _outer_border(ws, r, c1, r+4, c2, C_BORDER)

def _stat_card(ws, c1, c2, r, label, value, val_color):
    """4-row compact stat card"""
    _fill_range(ws, r, c1, r+3, c2, C_CARD)
    _mc(ws, r,   c1, r,   c2, label,
        font=_font(7, False, C_MUTED), align=_align())
    _mc(ws, r+1, c1, r+1, c2, value,
        font=_font(14, True, val_color), align=_align())
    _mc(ws, r+2, c1, r+2, c2, '')
    _mc(ws, r+3, c1, r+3, c2, '')
    _outer_border(ws, r, c1, r+3, c2, C_BORDER)

# â”€â”€ Data helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _index_name(symbol):
    sym = str(symbol).upper()
    for idx in ('BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY', 'NIFTY'):
        if sym.startswith(idx):
            return idx
    return sym

def _holding_minutes(entry_str, exit_str):
    if not entry_str or not exit_str:
        return 0
    fmts = ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M')
    et = xt = None
    for fmt in fmts:
        try:    et = datetime.strptime(str(entry_str).strip(), fmt); break
        except: pass
    for fmt in fmts:
        try:    xt = datetime.strptime(str(exit_str).strip(), fmt); break
        except: pass
    return max(0, int((xt - et).total_seconds() / 60)) if (et and xt) else 0

def _mins_to_hhmm(minutes):
    h, m = divmod(int(minutes), 60)
    return f'{h:02d}:{m:02d}'

def _safe_float(v):
    try:    return round(float(v or 0), 2)
    except: return 0.0

# â”€â”€ CSV load & process â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def load_trades():
    trades = []
    with open(CSV_PATH, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            if str(row.get('status', '')).upper() == 'OPEN':
                continue
            trades.append(row)
    trades.sort(key=lambda r: str(r.get('entry_time', '') or ''))
    return trades

def process_trades(raw):
    processed, cum = [], 0.0
    for r in raw:
        pnl = _safe_float(r.get('pnl'))
        cum = round(cum + pnl, 2)
        d   = str(r.get('direction', '') or '').upper()
        d   = 'BUY'  if d in ('BUY',  'BULLISH', 'LONG')  else \
              'SELL' if d in ('SELL', 'BEARISH', 'SHORT') else d
        processed.append({
            'date':         (r.get('date') or r.get('entry_time', ''))[:10],
            'index':        _index_name(r.get('symbol', '')),
            'direction':    d,
            'entry_price':  r.get('entry_price', ''),
            'exit_price':   r.get('exit_price', ''),
            'qty':          r.get('quantity', ''),
            'pnl':          pnl,
            'cum_pnl':      cum,
            'holding_mins': _holding_minutes(r.get('entry_time'), r.get('exit_time')),
            'rr':           _safe_float(r.get('rr_ratio')),
            'result':       str(r.get('result', '')).upper(),
            'entry_time':   r.get('entry_time', ''),
            'exit_time':    r.get('exit_time', ''),
        })
    return processed

def compute_kpis(trades):
    total        = len(trades)
    wins         = sum(1 for t in trades if t['result'] == 'WIN')
    losses       = sum(1 for t in trades if t['result'] == 'LOSS')
    total_pnl    = round(sum(t['pnl'] for t in trades), 2)
    win_rate     = round((wins / total * 100) if total else 0, 1)
    gross_profit = sum(t['pnl'] for t in trades if t['pnl'] > 0)
    gross_loss   = abs(sum(t['pnl'] for t in trades if t['pnl'] < 0))
    profit_factor = round(gross_profit / gross_loss, 2) if gross_loss else 0
    avg_rr   = round(sum(t['rr'] for t in trades) / total, 2) if total else 0
    avg_hold = round(sum(t['holding_mins'] for t in trades) / total) if total else 0

    index_pnl, index_cnt, index_wins = {}, {}, {}
    for t in trades:
        i = t['index']
        index_pnl[i]  = round(index_pnl.get(i, 0) + t['pnl'], 2)
        index_cnt[i]  = index_cnt.get(i, 0) + 1
        if t['result'] == 'WIN':
            index_wins[i] = index_wins.get(i, 0) + 1

    return {
        'total': total, 'wins': wins, 'losses': losses,
        'total_pnl': total_pnl, 'win_rate': win_rate,
        'profit_factor': profit_factor, 'avg_rr': avg_rr,
        'avg_hold': int(avg_hold),
        'best_trade':  max((t['pnl'] for t in trades), default=0),
        'worst_trade': min((t['pnl'] for t in trades), default=0),
        'gross_profit': round(gross_profit, 2),
        'gross_loss':   round(gross_loss, 2),
        'best_idx':  max(index_pnl, key=index_pnl.get) if index_pnl else 'N/A',
        'worst_idx': min(index_pnl, key=index_pnl.get) if index_pnl else 'N/A',
        'index_pnl': index_pnl, 'index_cnt': index_cnt, 'index_wins': index_wins,
    }

# â”€â”€ DASHBOARD sheet â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def build_dashboard(ws, kpis, trades):
    ws.sheet_view.showGridLines = False

    # Background fill â€” cover all visible content + chart area
    _fill_range(ws, 1, 1, 55, 19, C_BG)

    _setup_cols(ws)

    # â”€â”€ Row heights â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    _row_h(ws, {
        1:3,   2:5,   3:40,  4:18,  5:5,        # banner
        6:10,  7:4,   8:11,  9:38,  10:11, 11:4, # KPI cards
        12:10, 13:4,  14:11, 15:28, 16:4,        # stat cards
        17:14, 18:22, 19:22,                      # table header rows
        20:25, 21:25, 22:25, 23:25,               # index data rows (25pt each)
        24:12, 25:22,                             # spacer + chart labels
    })

    # â”€â”€ Banner â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    _fill_range(ws, 2, 2, 2, 18, C_BLUE)           # top accent line
    _fill_range(ws, 3, 1, 4, 19, C_BANNER)
    _fill_range(ws, 5, 2, 5, 18, C_BLUE)           # bottom accent line

    _mc(ws, 3, 2, 3, 18,
        'CB6 QUANTUM   Â·   ICT SILVER BULLET   Â·   TRADING DASHBOARD',
        font=_font(18, True, C_WHITE), align=_align())
    _mc(ws, 4, 2, 4, 18,
        f'Live Portfolio Analytics   Â·   NSE Index Futures   Â·   '
        f'Updated: {datetime.now().strftime("%d %b %Y  %H:%M")}',
        font=_font(9, False, C_MUTED, italic=True), align=_align())

    ws.freeze_panes = 'A6'

    # â”€â”€ KPI cards row 1 (rows 7-11) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    pf_str    = f'{kpis["profit_factor"]:.2f}' if kpis['profit_factor'] else 'N/A'
    hold_hhmm = _mins_to_hhmm(kpis['avg_hold'])

    kpi1 = [
        ('TOTAL CAPITAL',   'Rs 2,00,000',
         'Paper Trading Account',
         C_BLUE, C_WHITE),
        ('WIN RATE',        f'{kpis["win_rate"]}%',
         f'{kpis["wins"]}W  /  {kpis["losses"]}L  of  {kpis["total"]} trades',
         C_GREEN if kpis['win_rate'] >= 55 else C_AMBER,
         C_GREEN if kpis['win_rate'] >= 55 else C_AMBER),
        ('TOTAL P&L',       f'Rs {kpis["total_pnl"]:,.0f}',
         f'Gross Profit: Rs {kpis["gross_profit"]:,.0f}',
         C_GREEN if kpis['total_pnl'] >= 0 else C_RED,
         C_GREEN if kpis['total_pnl'] >= 0 else C_RED),
        ('PROFIT FACTOR',   pf_str,
         'Gross Profit / Gross Loss',
         C_GREEN if kpis['profit_factor'] >= 1.5 else C_AMBER,
         C_GREEN if kpis['profit_factor'] >= 1.5 else C_AMBER),
        ('AVG RISK:REWARD', f'1 : {kpis["avg_rr"]}',
         'Per completed trade',
         C_BLUE, C_WHITE),
        ('AVG HOLD TIME',   hold_hhmm,
         'HH:MM per trade',
         C_CYAN, C_WHITE),
    ]
    for (c1, c2), (lbl, val, sub, acc, vc) in zip(CARD_PAIRS, kpi1):
        _kpi_card(ws, c1, c2, 7, lbl, val, sub, acc, vc)

    # â”€â”€ Stat cards row 2 (rows 13-16) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    kpi2 = [
        ('TOTAL TRADES', str(kpis['total']),               C_WHITE),
        ('WINS',         str(kpis['wins']),                C_GREEN),
        ('LOSSES',       str(kpis['losses']),              C_RED),
        ('BEST TRADE',   f'Rs {kpis["best_trade"]:,.0f}',  C_GREEN),
        ('WORST TRADE',  f'Rs {kpis["worst_trade"]:,.0f}', C_RED),
        ('BEST INDEX',   kpis['best_idx'],
         INDEX_ACCENT.get(kpis['best_idx'], C_BLUE)),
    ]
    for (c1, c2), (lbl, val, vc) in zip(CARD_PAIRS, kpi2):
        _stat_card(ws, c1, c2, 13, lbl, val, vc)

    # â”€â”€ Index performance table (rows 18-23) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    _mc(ws, 18, 2, 18, 18, 'INDEX PERFORMANCE',
        fill=_fill(C_HDR), font=_font(10, True, C_BLUE), align=_align('left'))
    ws.cell(18, 2).border = Border(left=Side(border_style='thick', color=C_BLUE))

    # Column spans: INDEX(2-4) | TRADES(5-7) | WINS(8-10) | WIN%(11-13) | P&L(14-16) | AVG(17-18)
    tbl_cols = [(2,4),(5,7),(8,10),(11,13),(14,16),(17,18)]
    for (c1, c2), hdr in zip(tbl_cols,
                              ['INDEX','TRADES','WINS','WIN %','TOTAL P&L','AVG P&L']):
        _fill_range(ws, 19, c1, 19, c2, C_HDR)
        _mc(ws, 19, c1, 19, c2, hdr,
            font=_font(9, True, C_MUTED), align=_align())

    index_order = ['NIFTY', 'BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY']
    for off, idx in enumerate(index_order):
        row  = 20 + off
        cnt  = kpis['index_cnt'].get(idx, 0)
        wins = kpis['index_wins'].get(idx, 0)
        pnl  = kpis['index_pnl'].get(idx, 0.0)
        wr   = round(wins / cnt * 100, 1) if cnt else 0
        avgp = round(pnl / cnt, 0)         if cnt else 0
        rfc  = C_ROW_D if off % 2 == 0 else C_ROW_L
        pc   = C_GREEN if pnl >= 0 else C_RED
        wrc  = C_GREEN if wr  >= 55 else C_AMBER
        ic   = INDEX_ACCENT.get(idx, C_BLUE)

        for (c1, c2), (val, fc) in zip(tbl_cols, [
            (idx,               ic),
            (cnt,               C_WHITE),
            (wins,              C_GREEN if wins else C_MUTED),
            (f'{wr:.1f}%',      wrc),
            (f'Rs {pnl:,.0f}',  pc),
            (f'Rs {avgp:,.0f}', pc),
        ]):
            _fill_range(ws, row, c1, row, c2, rfc)
            _mc(ws, row, c1, row, c2, val,
                font=_font(9, c1 == 2, fc), align=_align())

    _outer_border(ws, 19, 2, 23, 18, C_BORDER)

    # â”€â”€ Chart section labels (row 25) â€” aligned with chart anchor columns â”€â”€â”€â”€â”€
    # Chart1 will be B26:J~45, Chart2 will be K26:R~45
    _mc(ws, 25, 2, 25, 9,                          # B25:I25 above chart1
        'CUMULATIVE P&L CURVE',
        fill=_fill(C_HDR), font=_font(9, True, C_BLUE), align=_align('left'))
    ws.cell(25, 2).border = Border(left=Side(border_style='thick', color=C_BLUE))

    _mc(ws, 25, 10, 25, 18,                        # J25:R25 above chart2
        'P&L BY INDEX',
        fill=_fill(C_HDR), font=_font(9, True, C_ORANGE), align=_align('left'))
    ws.cell(25, 10).border = Border(left=Side(border_style='thick', color=C_ORANGE))

    # â”€â”€ Chart data â€” narrow cols 23 and 25-26, plotVisOnly=False is critical â”€â”€
    # Equity curve data: col 23, rows 2 (header) + rows 3..N+2
    EQ_COL  = 23
    EQ_R0   = 2

    ws.cell(EQ_R0, EQ_COL).value = 'Equity Curve'
    for i, t in enumerate(trades, 1):
        ws.cell(EQ_R0 + i, EQ_COL).value = t['cum_pnl']

    # Bar chart data: col 25 (names) + col 26 (P&L), below equity data
    BAR_NM  = 25
    BAR_PNL = 26
    BAR_R0  = EQ_R0 + len(trades) + 3

    ws.cell(BAR_R0, BAR_NM).value  = 'Index'
    ws.cell(BAR_R0, BAR_PNL).value = 'P&L'
    for off, idx_name in enumerate(index_order, 1):
        ws.cell(BAR_R0 + off, BAR_NM).value  = idx_name
        ws.cell(BAR_R0 + off, BAR_PNL).value = kpis['index_pnl'].get(idx_name, 0)

    # â”€â”€ Chart 1: Cumulative P&L â€” AreaChart â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    chart1 = AreaChart()
    chart1.title         = None
    chart1.style         = 10
    chart1.grouping      = 'standard'
    chart1.y_axis.numFmt = '#,##0'
    chart1.y_axis.title  = 'P&L (Rs)'
    chart1.x_axis.title  = 'Trade #'
    chart1.height        = 11      # approx rows 26-45
    chart1.width         = 11      # approx cols B-J

    # CRITICAL: visible_cells_only=False ensures data in narrow columns renders
    chart1.visible_cells_only = False

    eq_ref = Reference(ws, min_col=EQ_COL,
                       min_row=EQ_R0, max_row=EQ_R0 + len(trades))
    chart1.add_data(eq_ref, titles_from_data=True)

    # 6-char RGB only â€” 8-char ARGB corrupts chart XML and triggers Excel repair
    chart1.series[0].graphicalProperties.solidFill          = '1E3A5F'
    chart1.series[0].graphicalProperties.line.solidFill     = '3B82F6'
    chart1.series[0].graphicalProperties.line.width         = 20000

    _transparent_plot_area(chart1)
    ws.add_chart(chart1, 'B26')

    # â”€â”€ Chart 2: Index P&L â€” BarChart, anchored 9 columns right of chart1 â”€â”€â”€â”€
    chart2 = BarChart()
    chart2.type          = 'col'
    chart2.title         = None
    chart2.style         = 10
    chart2.grouping      = 'clustered'
    chart2.y_axis.numFmt = '#,##0'
    chart2.y_axis.title  = 'P&L (Rs)'
    chart2.height        = 11
    chart2.width         = 11

    # CRITICAL: same fix
    chart2.visible_cells_only = False

    bar_vals = Reference(ws, min_col=BAR_PNL,
                         min_row=BAR_R0, max_row=BAR_R0 + len(index_order))
    bar_cats = Reference(ws, min_col=BAR_NM,
                         min_row=BAR_R0 + 1, max_row=BAR_R0 + len(index_order))
    chart2.add_data(bar_vals, titles_from_data=True)
    chart2.set_categories(bar_cats)

    # Per-bar colors: NIFTY=blue, BANKNIFTY=orange, FINNIFTY=purple, MIDCPNIFTY=cyan
    for i, color in enumerate(['3B82F6', 'F97316', 'A855F7', '06B6D4']):
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = color
        chart2.series[0].dPt.append(dp)

    _transparent_plot_area(chart2)
    ws.add_chart(chart2, 'K26')  # starts at col K â€” no overlap with chart1 at B

    # â”€â”€ Footer â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    _row_h(ws, {52: 22})
    _fill_range(ws, 52, 1, 52, 19, C_BANNER)
    _mc(ws, 52, 2, 52, 18,
        f'CB6 QUANTUM  Â·  Paper Trading  Â·  '
        f'Generated {datetime.now().strftime("%d %b %Y  %H:%M")}  Â·  NSE Index Futures Only',
        font=_font(8, False, C_MUTED, italic=True), align=_align())


# â”€â”€ TRADE_LOG sheet â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def build_trade_log(ws, trades):
    ws.sheet_view.showGridLines = False
    _fill_range(ws, 1, 1, len(trades) + 10, 12, C_BG)

    headers = [
        ('DATE',    11), ('INDEX', 14), ('DIR',  7),
        ('ENTRY',   12), ('EXIT',  12), ('QTY',  7),
        ('P&L',     13), ('HOLD',  9),  ('RR',   7),
        ('RESULT',  10),
    ]
    for j, (h, w) in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
        cell = ws.cell(1, j)
        cell.value     = h
        cell.fill      = _fill(C_HDR)
        cell.font      = _font(9, True, C_BLUE)
        cell.alignment = _align()
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    for i, t in enumerate(trades, 2):
        r   = t['result']
        rf  = C_ROW_D if i % 2 == 0 else C_ROW_L
        pc  = C_GREEN if t['pnl'] > 0 else C_RED if t['pnl'] < 0 else C_AMBER
        rc  = C_GREEN if r == 'WIN' else C_RED if r == 'LOSS' else C_AMBER
        ic  = INDEX_ACCENT.get(t['index'], C_WHITE)

        for j, (v, fc) in enumerate(zip([
            t['date'], t['index'], t['direction'],
            t['entry_price'], t['exit_price'], t['qty'],
            t['pnl'], _mins_to_hhmm(t['holding_mins']), t['rr'], r,
        ], [C_WHITE, ic, C_WHITE, C_WHITE, C_WHITE, C_WHITE, pc, C_MUTED, C_WHITE, rc]), 1):
            cell           = ws.cell(i, j)
            cell.value     = v
            cell.fill      = _fill(rf)
            cell.font      = _font(9, False, fc)
            cell.alignment = _align()
        ws.row_dimensions[i].height = 18

    last = len(trades) + 1
    tbl  = Table(displayName='TradeLog',
                 ref=f'A1:{get_column_letter(len(headers))}{last}')
    tbl.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2', showRowStripes=True,
        showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
    ws.add_table(tbl)


# â”€â”€ CALCS sheet (hidden) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def build_calcs(ws, kpis, trades):
    ws.sheet_state = 'hidden'
    rows = [
        ('CB6 CALCULATIONS', '', ''), ('', '', ''),
        ('KPI', 'VALUE', 'NOTES'),
        ('Total Trades',    kpis['total'],         ''),
        ('Wins',            kpis['wins'],          ''),
        ('Losses',          kpis['losses'],        ''),
        ('Win Rate %',      kpis['win_rate'],      ''),
        ('Total P&L',       kpis['total_pnl'],     'Rs'),
        ('Gross Profit',    kpis['gross_profit'],  'Rs'),
        ('Gross Loss',      kpis['gross_loss'],    'Rs'),
        ('Profit Factor',   kpis['profit_factor'], ''),
        ('Avg RR',          kpis['avg_rr'],        ''),
        ('Avg Hold (min)',  kpis['avg_hold'],       ''),
        ('Best Trade',      kpis['best_trade'],    'Rs'),
        ('Worst Trade',     kpis['worst_trade'],   'Rs'),
        ('', '', ''), ('INDEX BREAKDOWN', '', ''), ('Index', 'P&L', 'Trades'),
    ]
    for idx, pnl in kpis['index_pnl'].items():
        rows.append((idx, pnl, kpis['index_cnt'].get(idx, 0)))
    rows += [('', '', ''), ('EQUITY CURVE', '', ''), ('Trade#', 'PnL', 'Cum PnL')]
    for i, t in enumerate(trades, 1):
        rows.append((i, t['pnl'], t['cum_pnl']))

    for ri, row in enumerate(rows, 1):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci).value = val
    for col, w in [('A', 22), ('B', 16), ('C', 18)]:
        ws.column_dimensions[col].width = w


# â”€â”€ Main â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def main():
    print(f'Reading {CSV_PATH} ...')
    raw    = load_trades()
    trades = process_trades(raw)
    kpis   = compute_kpis(trades)
    print(f'  {len(trades)} trades  |  WR {kpis["win_rate"]}%  |  '
          f'P&L Rs {kpis["total_pnl"]:,.0f}  |  PF {kpis["profit_factor"]}')

    wb   = Workbook()
    ws_d = wb.active
    ws_d.title = 'DASHBOARD'
    ws_d.sheet_properties.tabColor = C_BLUE
    build_dashboard(ws_d, kpis, trades)

    ws_l = wb.create_sheet('TRADE_LOG')
    ws_l.sheet_properties.tabColor = C_GREEN
    build_trade_log(ws_l, trades)

    ws_c = wb.create_sheet('CALCS')
    build_calcs(ws_c, kpis, trades)

    wb.active = ws_d
    wb.save(OUT_PATH)
    print(f'Saved: {OUT_PATH}')


if __name__ == '__main__':
    main()

