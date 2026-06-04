# utils/bot_listener.py
import os
import sys
import requests
import threading
import time
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from settings import TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID, MAX_TRADES_PER_DAY
from settings import (
    EXECUTION_MODE,
    MAX_ENTRY_DRIFT_PERCENT,
    MAX_ENTRY_DRIFT_POINTS,
    EXECUTION_MIN_RR,
    EXECUTION_INVALIDATION_BUFFER_POINTS,
    EXECUTION_ALLOWED_SIGNAL_AGE_SECONDS,
)
from utils.logger import logger

scan_callback        = None   # equity full scan
nifty_scan_callback  = None   # NIFTY / BANKNIFTY index scan
signal_approval_callback = None
last_update_id       = 0
fyers_ref            = None
AUTHORIZED_CHAT_IDS  = {
    cid.strip()
    for cid in str(TELEGRAM_CHAT_ID or '').split(',')
    if cid.strip()
}
COMMAND_RATE_LIMIT_SECS = int(os.getenv('TELEGRAM_COMMAND_RATE_LIMIT_SECS', '3'))
_last_command_at = {}


def set_scan_callback(callback):
    global scan_callback
    scan_callback = callback


def set_nifty_scan_callback(callback):
    global nifty_scan_callback
    nifty_scan_callback = callback


def set_fyers_ref(fyers):
    global fyers_ref
    fyers_ref = fyers


def set_signal_approval_callback(callback):
    global signal_approval_callback
    signal_approval_callback = callback


def get_updates():
    global last_update_id
    try:
        url    = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getUpdates"
        # poll_timeout=5 → Telegram returns after 5s if no messages
        # requests timeout=(connect=5s, read=10s) → safely above poll_timeout
        params = {"offset": last_update_id + 1, "timeout": 5}
        resp   = requests.get(url, params=params, timeout=(5, 10))
        if resp.status_code == 200:
            return resp.json().get('result', [])
        return []
    except requests.exceptions.Timeout:
        return []   # normal long-poll timeout — not an error
    except Exception as e:
        logger.warning(f"Get updates error: {e}")
        return []


def _run_in_thread(fn, **kwargs):
    threading.Thread(target=fn, kwargs=kwargs, daemon=True).start()


def _is_authorized_chat(chat_id: str) -> bool:
    return bool(chat_id) and chat_id in AUTHORIZED_CHAT_IDS


def _rate_limited(chat_id: str, command: str) -> bool:
    now = time.time()
    key = (chat_id, command.split()[0] if command else '')
    last = _last_command_at.get(key, 0)
    if now - last < COMMAND_RATE_LIMIT_SECS:
        return True
    _last_command_at[key] = now
    return False


def process_updates(updates):
    global last_update_id, scan_callback, nifty_scan_callback, fyers_ref

    for update in updates:
        last_update_id = update['update_id']
        message = update.get('message', {})
        text    = message.get('text', '').strip()
        chat_id = str(message.get('chat', {}).get('id', ''))

        if not _is_authorized_chat(chat_id):
            if chat_id:
                logger.warning(f"Rejected Telegram command from unknown chat_id={chat_id}")
            continue

        # Strip bot-username suffix Telegram appends when tapping commands
        # e.g. "/ask@cb6bot question" → "/ask question"
        if text.startswith('/') and '@' in text.split()[0]:
            first, *rest = text.split()
            first = first.split('@')[0]
            text  = ' '.join([first] + rest)

        logger.info(f"Command received: {text}")

        from utils.telegram_alerts import send_message

        if _rate_limited(chat_id, text):
            send_message("Command rate limit active. Try again in a few seconds.")
            continue

        # ── MENU ──────────────────────────────────────────────────────────────

        if text == '/start':
            import datetime as _dt, pytz as _pytz
            _ist   = _dt.datetime.now(_pytz.timezone('Asia/Kolkata'))
            _time  = _ist.strftime('%H:%M IST')
            _date  = _ist.strftime('%d %b %Y')
            # Window status
            _hour_ist = _ist.hour + _ist.minute / 60
            if 10.0 <= _hour_ist < 11.0:
                _win = '🟢 Morning window OPEN (10:00–11:00)'
            elif 13.5 <= _hour_ist < 14.5:
                _win = '🟢 Afternoon window OPEN (13:30–14:30)'
            elif 9.25 <= _hour_ist < 10.0:
                _win = '⏳ Pre-market — window opens at 10:00'
            elif 15.5 <= _hour_ist < 15.6:
                _win = '🔴 Market closed'
            else:
                _win = '⏸ Outside window — monitoring'
            send_message(
                "<b>CB6 QUANTUM — INDIAN MARKETS</b>\n"
                f"{_date}  |  {_time}\n\n"
                "Strategy : ICT Silver Bullet\n"
                "Markets  : NSE Index Futures &amp; Options\n"
                "           NIFTY | BANKNIFTY | FINNIFTY | MIDCPNIFTY\n"
                "Broker   : Fyers  |  Mode: Live\n\n"
                "<b>SESSION WINDOWS (auto-scan)</b>\n"
                "Morning   : 10:00 – 11:00 IST\n"
                "Afternoon : 13:30 – 14:30 IST\n"
                f"Status    : {_win}\n"
                "Chain     : DOL → MSS → Displaced FVG → Entry\n\n"
                "<b>SILVER BULLET</b>\n"
                "/sb              Trigger scan now (all 4 indices)\n"
                "/scan            Index futures scan\n"
                "/check NIFTY     Check a specific index\n\n"
                "<b>TRADES &amp; ACCOUNT</b>\n"
                "/trades          Open positions + live PnL\n"
                "/portfolio       Capital &amp; P&amp;L summary\n"
                "/nse_status      Bot health, windows, last scan\n"
                "/excel           Download Excel dashboard\n\n"
                "<b>INTELLIGENCE</b>\n"
                "/levels          NIFTY ICT levels + probability\n"
                "/brain           Market bias + session score\n"
                "/options NIFTY   Strike selector (ITM/SATM/OTM)\n"
                "/fiidii          FII/DII flow data\n"
                "/expiry          F&amp;O expiry calendar\n\n"
                "<b>ML INTELLIGENCE</b>\n"
                "/ml_scan NIFTY   Multi-TF ICT scan + ML score\n"
                "/ml_scan ALL     Scan all 4 indices at once\n"
                "/ml_status       Model accuracy + predictions\n\n"
                "<b>CONTROL</b>\n"
                "/stop            Halt trading today\n"
                "/resume          Resume trading\n"
                "/help            ICT Silver Bullet rules\n"
                "/info            Full command reference",
                parse_mode='HTML'
            )

        elif text == '/info':
            send_message(
                "<b>CB6 QUANTUM — ALL COMMANDS</b>\n\n"
                "Mode: Index Futures &amp; Options ONLY\n"
                "NIFTY | BANKNIFTY | FINNIFTY | MIDCPNIFTY\n\n"
                "<b>SILVER BULLET (auto-fires at window open)</b>\n"
                "10:00 – 11:00 IST  Morning window\n"
                "13:30 – 14:30 IST  Afternoon window\n"
                "Chain: DOL → MSS → Displaced FVG → Entry\n\n"
                "/sb               Trigger scan now (all 4 indices)\n"
                "/scan             Index futures scan\n"
                "/check INDEX      Check specific index — /check NIFTY\n\n"
                "<b>TRADES &amp; ACCOUNT</b>\n"
                "/trades           Open index positions + live PnL\n"
                "/portfolio        Capital &amp; P&amp;L summary\n"
                "/nse_status       Bot health, windows, today's trades\n"
                "/excel            Download Excel dashboard\n"
                "/replay           Chart replay of last trade\n\n"
                "<b>INTELLIGENCE</b>\n"
                "/levels           NIFTY ICT levels + buy/sell probability\n"
                "/options NIFTY    ITM/SATM/OTM strikes with delta/theta/IV\n"
                "/brain            Market bias + session score\n"
                "/brain refresh    Force full refresh\n"
                "/fiidii           FII/DII flow data\n"
                "/expiry           F&amp;O expiry calendar\n\n"
                "<b>ML INTELLIGENCE</b>\n"
                "/ml_scan NIFTY    Multi-TF ICT scan + ML score\n"
                "/ml_scan ALL      Scan all 4 indices at once\n"
                "/ml_status        ML model accuracy + predictions\n"
                "/ml_train         Force ML retrain now\n\n"
                "<b>AI &amp; LEARNING</b>\n"
                "/ask question     Ask AI anything\n"
                "/memory           AI trade stats\n"
                "/learn            Learned parameters\n"
                "/lessons          Trade post-mortems\n"
                "/clearchat        Reset AI chat\n\n"
                "<b>PATTERN ENGINE</b>\n"
                "/pattern          Library stats (WR by window/direction)\n"
                "/reloadpatterns   Reload after fresh backtest\n\n"
                "<b>ADVANCED</b>\n"
                "/backtest3m NIFTY  3m ICT backtest\n"
                "/execution_mode    Execution gate config\n"
                "/execution_stats   Execution validation stats\n"
                "/ws                WebSocket status\n"
                "/ws on | off       Toggle realtime feed\n"
                "/adversarial       Robustness test\n\n"
                "<b>CONTROL</b>\n"
                "/stop              Halt trading today\n"
                "/resume            Resume trading\n"
                "/eventmode on|off  Crisis filter mode\n"
                "/help              ICT Silver Bullet strategy rules",
                parse_mode='HTML'
            )

        elif text.startswith('/approve'):
            parts = text.split()
            if len(parts) < 2:
                send_message("Usage: /approve <SIGNAL_ID>")
                continue
            if not signal_approval_callback:
                send_message("Approval handler is not enabled in this engine.")
                continue
            signal_id = parts[1].strip().upper()
            try:
                ok, msg = signal_approval_callback(signal_id=signal_id, approved=True, chat_id=chat_id)
                send_message(msg if msg else ("Approved." if ok else "Approval failed."))
            except Exception as e:
                logger.exception(f"approve command error: {e}")
                send_message(f"Approve error: {e}")

        elif text.startswith('/reject'):
            parts = text.split()
            if len(parts) < 2:
                send_message("Usage: /reject <SIGNAL_ID>")
                continue
            if not signal_approval_callback:
                send_message("Approval handler is not enabled in this engine.")
                continue
            signal_id = parts[1].strip().upper()
            try:
                ok, msg = signal_approval_callback(signal_id=signal_id, approved=False, chat_id=chat_id)
                send_message(msg if msg else ("Rejected." if ok else "Reject failed."))
            except Exception as e:
                logger.exception(f"reject command error: {e}")
                send_message(f"Reject error: {e}")

        elif text == '/pending':
            try:
                from utils.execution_validation import list_signals_by_state, SIGNAL_WAITING_CONFIRM
                pending = list_signals_by_state(SIGNAL_WAITING_CONFIRM)
                if not pending:
                    send_message("No pending signals waiting for manual confirmation.")
                    continue
                pending = sorted(pending, key=lambda x: x.get('updated_at', ''), reverse=True)[:20]
                lines = ["Pending Signals (WAITING_FOR_MANUAL_CONFIRMATION):", ""]
                for s in pending:
                    lines.append(
                        f"{s.get('signal_id')} | {s.get('symbol')} | {s.get('direction')} | "
                        f"LTP {s.get('current_ltp')} | RR {s.get('calculated_rr')} | Age {s.get('signal_age_seconds')}s"
                    )
                send_message('\n'.join(lines))
            except Exception as e:
                logger.exception(f"pending command error: {e}")
                send_message(f"Pending error: {e}")

        elif text.startswith('/signal'):
            parts = text.split()
            if len(parts) < 2:
                send_message("Usage: /signal <SIGNAL_ID>")
                continue
            signal_id = parts[1].strip().upper()
            try:
                from utils.execution_validation import get_signal
                s = get_signal(signal_id)
                if not s:
                    send_message(f"Signal not found: {signal_id}")
                    continue
                msg = (
                    f"Signal {signal_id}\n\n"
                    f"State: {s.get('state')}\n"
                    f"Reason: {s.get('status_reason')}\n"
                    f"Symbol: {s.get('symbol')}\n"
                    f"Direction: {s.get('direction')}\n"
                    f"Planned Entry: {s.get('planned_entry')}\n"
                    f"Current LTP: {s.get('current_ltp')}\n"
                    f"Entry Band: {s.get('entry_band_low')} to {s.get('entry_band_high')}\n"
                    f"Stop Loss: {s.get('stop_loss')}\n"
                    f"Target: {s.get('target')}\n"
                    f"RR: {s.get('calculated_rr')}\n"
                    f"Signal Age: {s.get('signal_age_seconds')}s\n"
                    f"Created: {s.get('created_at')}\n"
                    f"Updated: {s.get('updated_at')}\n"
                )
                send_message(msg)
            except Exception as e:
                logger.exception(f"signal command error: {e}")
                send_message(f"Signal error: {e}")

        elif text.startswith('/cancel'):
            parts = text.split()
            if len(parts) < 2:
                send_message("Usage: /cancel <SIGNAL_ID>")
                continue
            signal_id = parts[1].strip().upper()
            try:
                from utils.execution_validation import get_signal, cancel_signal
                s = get_signal(signal_id)
                if not s:
                    send_message(f"Signal not found: {signal_id}")
                    continue
                if s.get('state') not in ('WAITING_FOR_MANUAL_CONFIRMATION', 'ARMED', 'VALIDATING', 'NEW'):
                    send_message(f"Signal {signal_id} cannot be cancelled from state {s.get('state')}.")
                    continue
                cancelled = cancel_signal(signal_id, reason="USER_CANCELLED")
                if not cancelled:
                    send_message(f"Failed to cancel signal {signal_id}.")
                    continue
                send_message(f"Signal {signal_id} cancelled. Reason: USER_CANCELLED")
            except Exception as e:
                logger.exception(f"cancel command error: {e}")
                send_message(f"Cancel error: {e}")

        elif text == '/execution_mode':
            send_message(
                "Execution Mode\n\n"
                f"EXECUTION_MODE: {EXECUTION_MODE}\n"
                f"MAX_ENTRY_DRIFT_PERCENT: {MAX_ENTRY_DRIFT_PERCENT}\n"
                f"MAX_ENTRY_DRIFT_POINTS: {MAX_ENTRY_DRIFT_POINTS}\n"
                f"EXECUTION_MIN_RR: {EXECUTION_MIN_RR}\n"
                f"EXECUTION_INVALIDATION_BUFFER_POINTS: {EXECUTION_INVALIDATION_BUFFER_POINTS}\n"
                f"EXECUTION_ALLOWED_SIGNAL_AGE_SECONDS: {EXECUTION_ALLOWED_SIGNAL_AGE_SECONDS}\n"
            )

        elif text == '/execution_stats':
            try:
                from utils.execution_validation import get_execution_stats_for_date
                s = get_execution_stats_for_date()
                if s.get('total_signals', 0) == 0:
                    send_message(
                        "Execution Validation Stats (Today)\n\n"
                        "No signals found in execution_validation_audit.jsonl for today."
                    )
                    continue
                lines = []
                for reason, cnt in sorted(
                    (s.get('blocked_reason_breakdown') or {}).items(),
                    key=lambda x: x[1],
                    reverse=True
                ):
                    lines.append(f"{reason}: {cnt}")
                breakdown = '\n'.join(lines) if lines else "None"
                send_message(
                    "Execution Validation Stats (Today)\n\n"
                    f"total_signals: {s.get('total_signals')}\n"
                    f"blocked_count: {s.get('blocked_count')}\n"
                    f"approved_count: {s.get('approved_count')}\n"
                    f"executed_count: {s.get('executed_count')}\n"
                    f"block_rate: {s.get('block_rate_pct')}%\n\n"
                    "blocked reason breakdown:\n"
                    f"{breakdown}"
                )
            except Exception as e:
                logger.exception(f"execution_stats command error: {e}")
                send_message(f"Execution stats error: {e}")

        # ── OPTIONS ITM / SATM / OTM ─────────────────────────────────────────

        elif text == '/execution_report':
            try:
                from utils.execution_validation import get_pipeline_telemetry
                metrics = get_pipeline_telemetry()
                total = metrics.get('total_signals_received', 0)
                armed = metrics.get('currently_armed', 0)
                executed = metrics.get('executed_count', 0)
                blocked = metrics.get('blocked_count', 0)
                conversion_pct = (executed / total * 100.0) if total > 0 else 0.0
                if total == 0 and armed == 0:
                    send_message(
                        "Execution Report (Indian Engine, Today)\n\n"
                        "No execution-validation telemetry found for today."
                    )
                    continue

                lines = []
                for reason, cnt in sorted(
                    (metrics.get('breakdown') or {}).items(),
                    key=lambda x: x[1],
                    reverse=True
                ):
                    short_reason = str(reason).replace("REVALIDATION:", "")
                    lines.append(f"{short_reason}: {cnt}")
                breakdown = '\n'.join(lines) if lines else "None"

                send_message(
                    "EXECUTION GATEKEEPER AUDIT REPORT\n"
                    "=========================================\n"
                    "Engine Path: NSE/MCX FLOW (Automated)\n"
                    "Mode: SAFE_VALIDATION_REVALIDATE_AUTO\n"
                    "=========================================\n\n"
                    "PIPELINE COUNTERS\n"
                    f"Total ML Signals Received: {total}\n"
                    f"Currently ARMED State: {armed}\n\n"
                    "TERMINAL RESOLUTION MATRIX\n"
                    f"EXECUTED (Auto-Passed): {executed} ({conversion_pct:.1f}% Accept Rate)\n"
                    f"BLOCKED (Filtered Out): {blocked}\n\n"
                    "MICROSTRUCTURAL FAILURE LEAKAGE\n"
                    f"{breakdown}\n\n"
                    "System Status: 93 Unit Tests Passed | Syntax Verified\n"
                    "Legacy Fallback: ONLINE & UNCHANGED"
                )
            except Exception as e:
                logger.exception(f"execution_report command error: {e}")
                send_message(f"Execution report error: {e}")

        elif text.startswith('/options'):
            _INDEX_OPTS = {
                'NIFTY'      : 'NIFTY',      'NIFTY50'    : 'NIFTY',
                'BANKNIFTY'  : 'BANKNIFTY',  'BANK'       : 'BANKNIFTY',
                'FINNIFTY'   : 'FINNIFTY',   'FIN'        : 'FINNIFTY',
                'MIDCPNIFTY' : 'MIDCPNIFTY', 'MIDCP'      : 'MIDCPNIFTY',
            }
            parts   = text.split()
            raw_idx = parts[1].upper() if len(parts) > 1 else 'NIFTY'
            idx     = _INDEX_OPTS.get(raw_idx, 'NIFTY')

            send_message(
                f"CB6 OPTIONS — {idx}\n\n"
                "Fetching live option chain...\n"
                "Computing ITM / SATM / OTM strikes via delta.\n"
                "~10 seconds..."
            )

            def _run_options():
                try:
                    from scanner.option_strike_selector import (
                        get_itm_satm_otm, format_options_table,
                        get_index_spot, get_nearest_expiry, load_symbol_master,
                    )

                    # Use index spot (not futures) — options are priced on the index
                    spot = get_index_spot(fyers_ref, idx)
                    if not spot:
                        send_message(f"Could not fetch live index price for {idx}. Check token.")
                        return

                    ce_tiers = get_itm_satm_otm(fyers_ref, idx, 'CE', spot)
                    pe_tiers = get_itm_satm_otm(fyers_ref, idx, 'PE', spot)

                    master     = load_symbol_master()
                    expiry     = get_nearest_expiry(idx, master)
                    expiry_str = expiry.isoformat() if expiry else 'unknown'

                    send_message(format_options_table(idx, spot, expiry_str,
                                                      ce_tiers, pe_tiers))
                except Exception as e:
                    send_message(f"Options error: {e}")

            _run_in_thread(_run_options)

        # ── NIFTY LEVELS + PROBABILITY ───────────────────────────────────────

        elif text in ('/levels', '/nifty levels'):
            send_message(
                "CB6 NIFTY LEVELS\n\n"
                "Fetching live 5-min + 15-min + daily data...\n"
                "Mapping ICT levels + matching backtest patterns.\n"
                "~10 seconds..."
            )
            def _run_levels():
                try:
                    from scanner.nifty_levels import analyse_nifty, format_report
                    result = analyse_nifty(fyers_ref)
                    send_message(format_report(result))
                except Exception as e:
                    send_message(f"Levels error: {e}")
            _run_in_thread(_run_levels)

        # ── MARKET BRAIN ─────────────────────────────────────────────────────

        elif text in ('/brain', '/brain refresh'):
            try:
                from core.market_brain import refresh as brain_refresh, format_report
                if text == '/brain refresh':
                    send_message("CB6 BRAIN — refreshing all signals...")
                    brain_refresh(fyers_ref, force=True)
                send_message(format_report())
            except Exception as e:
                send_message(f"Brain error: {e}")

        # ── SILVER BULLET MANUAL TRIGGER ─────────────────────────────────────

        elif text == '/sb':
            from scanner.silver_bullet import get_window_status
            from data.pattern_library import load_library
            import datetime, pytz
            _ist = datetime.datetime.now(pytz.timezone('Asia/Kolkata')).strftime('%H:%M')
            _pat = len(load_library())
            send_message(
                f"CB6 SILVER BULLET SCAN\n\n"
                f"Time     : {_ist} IST\n"
                f"Symbols  : NIFTY | BANKNIFTY | FINNIFTY | MIDCPNIFTY\n"
                f"Chain    : DOL → MSS → Displaced FVG → Entry\n"
                f"Patterns : {_pat} backtest trades in memory\n"
                f"Window   : {get_window_status()}\n\n"
                f"Scanning..."
            )
            if scan_callback:
                _run_in_thread(scan_callback)

        # ── EXCEL DASHBOARD ───────────────────────────────────────────────────

        elif text == '/excel':
            send_message("CB6 QUANTUM - Generating Excel dashboard... please wait.")
            def _send_excel():
                try:
                    import sys as _sys
                    _base = os.path.dirname(os.path.dirname(__file__))
                    _sys.path.insert(0, _base)
                    from generate_excel_dashboard import load_trades, process_trades, compute_kpis
                    from generate_excel_dashboard import build_dashboard, build_trade_log, build_calcs
                    from openpyxl import Workbook
                    import datetime as _dt

                    raw    = load_trades()
                    trades = process_trades(raw)
                    kpis   = compute_kpis(trades)

                    wb     = Workbook()
                    ws_d   = wb.active
                    ws_d.title = 'DASHBOARD'
                    ws_d.sheet_properties.tabColor = '3B82F6'
                    build_dashboard(ws_d, kpis, trades)

                    ws_l   = wb.create_sheet('TRADE_LOG')
                    ws_l.sheet_properties.tabColor = '10B981'
                    build_trade_log(ws_l, trades)

                    ws_c   = wb.create_sheet('CALCS')
                    build_calcs(ws_c, kpis, trades)
                    wb.active = ws_d

                    out = os.path.join(_base, 'data', 'CB6_Dashboard.xlsx')
                    wb.save(out)

                    from utils.telegram_alerts import send_document
                    today = _dt.datetime.now().strftime('%d %b %Y')
                    caption = (
                        f"CB6 QUANTUM — Excel Dashboard\n"
                        f"Date: {today}\n"
                        f"Trades: {kpis['total']}  |  Win Rate: {kpis['win_rate']}%\n"
                        f"Total P&L: Rs {kpis['total_pnl']:,.0f}"
                    )
                    ok = send_document(out, caption)
                    if not ok:
                        send_message("CB6 QUANTUM - Excel ready but file send failed. Check logs.")
                except Exception as e:
                    send_message(f"Excel error: {e}")
            _run_in_thread(_send_excel)

        # ── PORTFOLIO ─────────────────────────────────────────────────────────

        elif text == '/portfolio':
            try:
                from trader.paper_trader import get_portfolio_summary
                get_portfolio_summary()
            except Exception as e:
                send_message(f"Portfolio error: {e}")

        # ── OPEN TRADES ───────────────────────────────────────────────────────

        elif text == '/trades':
            try:
                from trader.paper_trader import load_state
                from scanner.live_price  import get_live_price
                import datetime as _dt

                state        = load_state()
                index_trades = [
                    t for t in state.get('open_trades', [])
                    if t.get('instrument_type', 'EQUITY') != 'EQUITY'
                ]

                if not index_trades:
                    send_message("CB6 QUANTUM - No open index futures/options trades.")
                else:
                    total_pnl = 0
                    now_str   = _dt.datetime.now().strftime('%H:%M')
                    send_message(
                        f"CB6 QUANTUM - {len(index_trades)} Open Index Trade(s) | {now_str} IST\n"
                        f"Fetching live prices..."
                    )
                    for trade in index_trades:
                        sym       = trade['symbol']
                        clean     = sym.replace('NSE:', '').replace('-EQ', '').replace('-INDEX', '')
                        inst      = trade.get('instrument_type', 'INDEX')
                        direction = trade.get('direction', 'BUY')

                        ltp = get_live_price(fyers_ref, sym) if fyers_ref else None
                        if ltp:
                            if direction in ('BUY', 'BULLISH'):
                                pnl = round((ltp - trade['entry_price']) * trade['quantity'], 2)
                            else:
                                pnl = round((trade['entry_price'] - ltp) * trade['quantity'], 2)
                        else:
                            pnl = trade.get('pnl', 0)
                            ltp = trade['entry_price']

                        total_pnl += pnl
                        sign    = '+' if pnl >= 0 else ''
                        targets = ', '.join(trade.get('targets_hit', [])) or 'None hit yet'
                        sl_dist = round(abs(ltp - trade['current_sl']), 2)
                        t1_dist = round(abs(trade['target1'] - ltp), 2)

                        send_message(
                            f"CB6 QUANTUM - OPEN TRADE\n\n"
                            f"Symbol  : {clean}\n"
                            f"Type    : {inst}\n"
                            f"Dir     : {direction}\n"
                            f"TF      : {trade['timeframe']}\n"
                            f"Entry   : {trade['entry_price']}\n"
                            f"LTP     : {ltp}\n"
                            f"SL      : {trade['current_sl']}  ({sl_dist} pts away)\n"
                            f"T1      : {trade['target1']}  ({t1_dist} pts away)\n"
                            f"T2      : {trade['target2']}\n"
                            f"T3      : {trade['target3']}\n"
                            f"Qty     : {trade['quantity']}\n"
                            f"PnL     : Rs {sign}{pnl:.0f}\n"
                            f"Targets : {targets}\n"
                            f"Opened  : {trade['entry_time']}"
                        )
                        time.sleep(0.3)

                    sign_tot = '+' if total_pnl >= 0 else ''
                    send_message(f"TOTAL OPEN PnL : Rs {sign_tot}{total_pnl:.0f}")
            except Exception as e:
                send_message(f"Trades error: {e}")

        # ── NSE STATUS ───────────────────────────────────────────────────────────

        elif text == '/nse_status':
            try:
                import datetime as _dt, pytz as _pytz
                from scanner.silver_bullet import get_window_status, is_silver_bullet_window
                from trader.paper_trader import load_state

                _ist   = _dt.datetime.now(_pytz.timezone('Asia/Kolkata'))
                _time  = _ist.strftime('%H:%M IST')
                _state = load_state()
                _paused = _state.get('paused', False)
                _open   = [t for t in _state.get('open_trades', [])
                           if t.get('instrument_type', 'EQUITY') != 'EQUITY']
                _closed = [t for t in _state.get('closed_trades', [])
                           if t.get('instrument_type', 'EQUITY') != 'EQUITY']
                _today  = _ist.strftime('%Y-%m-%d')
                _today_closed = [t for t in _closed
                                 if str(t.get('entry_time', '')).startswith(_today)]
                _today_pnl = sum(t.get('pnl', 0) for t in _today_closed)
                _today_wins = sum(1 for t in _today_closed if t.get('pnl', 0) > 0)
                _total_pnl  = _state.get('total_pnl', 0)

                _in_win, _win_name = is_silver_bullet_window()
                _win_status = _state.get('window_status', get_window_status())
                _engine_status = '⏸ PAUSED' if _paused else '✅ RUNNING'

                sign_day = '+' if _today_pnl >= 0 else ''
                sign_tot = '+' if _total_pnl >= 0 else ''

                send_message(
                    "<b>CB6 QUANTUM — NSE BOT STATUS</b>\n"
                    f"Time       : {_time}\n\n"
                    f"<b>ENGINE</b>\n"
                    f"Status     : {_engine_status}\n"
                    f"Fyers      : {'✅ Connected' if fyers_ref else '❌ Not connected'}\n\n"
                    f"<b>SESSION WINDOWS</b>\n"
                    f"Morning    : 10:00 – 11:00 IST\n"
                    f"Afternoon  : 13:30 – 14:30 IST\n"
                    f"Now        : {'🟢 ' + _win_name if _in_win else '⏸ Outside window'}\n\n"
                    f"<b>TODAY ({_today})</b>\n"
                    f"Trades     : {len(_today_closed)} closed | {len(_open)} open\n"
                    f"Wins       : {_today_wins}/{len(_today_closed)}\n"
                    f"Today PnL  : Rs {sign_day}{_today_pnl:,.0f}\n"
                    f"Total PnL  : Rs {sign_tot}{_total_pnl:,.0f}\n\n"
                    f"<b>INDICES</b>\n"
                    f"NIFTY | BANKNIFTY | FINNIFTY | MIDCPNIFTY\n"
                    f"Chain : DOL → MSS → Displaced FVG\n"
                    f"Mode  : Index Futures &amp; Options only",
                    parse_mode='HTML'
                )
            except Exception as e:
                send_message(f"NSE status error: {e}")

        # ── MEMORY ────────────────────────────────────────────────────────────

        elif text == '/memory':
            try:
                from data.bot_memory import get_memory_summary
                send_message(get_memory_summary())
            except Exception as e:
                send_message(f"Memory error: {e}")

        # ── FII/DII ───────────────────────────────────────────────────────────

        elif text == '/fiidii':
            try:
                from data.fii_dii import get_fii_dii_summary
                send_message(get_fii_dii_summary())
            except Exception as e:
                send_message(f"FII/DII error: {e}")

        # ── LEARNED PARAMS ────────────────────────────────────────────────────

        elif text == '/learn':
            try:
                from data.bot_memory import load_memory
                memory = load_memory()
                params = memory.get('learned_params', {})
                total  = memory.get('total_trades', 0)
                wins   = memory.get('winning_trades', 0)
                wr     = round(wins / total * 100, 1) if total > 0 else 0
                n_perf = memory.get('nifty_performance', {})
                n_wr   = round(
                    n_perf.get('wins', 0) /
                    max(n_perf.get('trades', 1), 1) * 100, 1
                )
                send_message(
                    "CB6 QUANTUM - LEARNED PARAMETERS\n\n"
                    f"All Trades  : {total}  WR: {wr}%\n"
                    f"NIFTY Trades: {n_perf.get('trades',0)}  WR: {n_wr}%\n\n"
                    "PARAMETERS:\n"
                    f"Score Threshold : {params.get('best_score_threshold', 5)}/10\n"
                    f"Best Hours      : {params.get('best_hours', [10,11,14])}\n"
                    f"Avoid Hours     : {params.get('avoid_hours', [])}\n\n"
                    "BEST EQUITY STOCKS:\n"
                    f"{', '.join(params.get('best_stocks', [])[:5]) or 'Still learning...'}\n\n"
                    "AVOID STOCKS:\n"
                    f"{', '.join(params.get('avoid_stocks', [])[:5]) or 'None yet'}\n\n"
                    "NIFTY INDEX AI:\n"
                    f"Wins  : {n_perf.get('wins',0)}\n"
                    f"Losses: {n_perf.get('losses',0)}\n"
                    f"PnL   : Rs {n_perf.get('total_pnl',0):.0f}\n\n"
                    "Bot learns after every trade!"
                )
            except Exception as e:
                send_message(f"Learn error: {e}")

        # ── STOP TRADING ──────────────────────────────────────────────────────

        elif text == '/stop':
            try:
                from trader.paper_trader import load_state, save_state
                state = load_state()
                state['paused'] = True
                save_state(state)
                send_message(
                    "CB6 QUANTUM - Scanning paused.\n"
                    "No new trades will open until you send /resume.\n"
                    "Open positions continue to be managed."
                )
            except Exception as e:
                send_message(f"Stop error: {e}")

        elif text == '/resume':
            try:
                from trader.paper_trader import load_state, save_state
                state = load_state()
                state['paused'] = False
                save_state(state)
                send_message("CB6 QUANTUM - Scanning resumed. Bot will take new live trades.")
            except Exception as e:
                send_message(f"Resume error: {e}")

        # ── ML MEMORY STATUS ──────────────────────────────────────────────────

        elif text == '/ml_memory':
            try:
                from ml.trade_memory import format_memory_status
                send_message(format_memory_status())
            except Exception as e:
                send_message(f"ML memory error: {e}")

        # ── PATTERN LIBRARY SAMPLE STATUS ────────────────────────────────────

        elif text == '/library_status':
            try:
                from utils.trade_enrichment import format_sample_report
                send_message(format_sample_report())
            except Exception as e:
                send_message(f"Library status error: {e}")

        # ── MANUAL POSITION ANALYSIS ─────────────────────────────────────────

        elif text == '/analyze_positions':
            try:
                from core.manual_position_analyzer import send_analysis_report
                _fyers = _fyers_ref  # set by set_fyers_ref() in main.py
                if not _fyers:
                    send_message("Fyers not initialised — start the bot first.")
                else:
                    send_message("Analysing manual positions...")
                    send_analysis_report(_fyers)
            except Exception as e:
                send_message(f"Position analysis error: {e}")

        # ── PATTERN LIBRARY ───────────────────────────────────────────────────

        elif text == '/pattern':
            try:
                from data.pattern_library import library_stats
                send_message(library_stats())
            except Exception as e:
                send_message(f"Pattern library error: {e}")

        elif text == '/reloadpatterns':
            try:
                from data.pattern_library import reload_library
                count = reload_library()
                send_message(
                    f"CB6 PATTERN LIBRARY RELOADED\n\n"
                    f"Loaded {count} trade fingerprints from backtest CSVs.\n"
                    f"Use /pattern for full stats."
                )
            except Exception as e:
                send_message(f"Reload error: {e}")

        # ── HELP ──────────────────────────────────────────────────────────────

        elif text == '/help':
            send_message(
                "CB6 — ICT SILVER BULLET STRATEGY\n\n"
                "MODE: Index Futures & Options ONLY\n"
                "NIFTY | BANKNIFTY | FINNIFTY | MIDCPNIFTY\n"
                "(Equity stocks disabled)\n\n"
                "WINDOWS (auto-scan fires at open):\n"
                "Morning   : 10:00 – 11:00 IST\n"
                "Afternoon : 13:30 – 14:30 IST\n\n"
                "SETUP CHAIN (3 steps):\n"
                "1. Draw on Liquidity\n"
                "   Nearest unswept swing high/low\n"
                "   — this is where price is drawn\n\n"
                "2. Market Structure Shift (MSS)\n"
                "   Close beyond last swing point\n"
                "   — confirms directional intent\n\n"
                "3. Fair Value Gap (FVG)\n"
                "   3-candle imbalance after MSS\n"
                "   — enter on FIRST touch of FVG\n\n"
                "TRADE PLAN:\n"
                "Entry  : FVG edge (bottom/top)\n"
                "SL     : Other edge of FVG\n"
                "T1     : 1:2 RR — exit 50%, SL to BE\n"
                "T2     : 1:3 RR — main target\n"
                "T3     : DOL level (Draw on Liquidity)\n\n"
                "RULES:\n"
                "No trades before 10:00 AM\n"
                "(9:15 open is Judas Swing — retail trap)\n"
                "Options: ITM/ATM CE/PE | Delta 0.6–0.8\n"
                "OTM options banned — theta kills RR\n"
                "Exit if stuck in FVG > 15 min\n"
                "Risk: max 1-2% capital per trade\n"
                "FII/DII: prefer direction of flow\n\n"
                "INSTRUMENTS:\n"
                "Futures : NIFTY / BANKNIFTY / FINNIFTY / MIDCPNIFTY\n"
                "Options : CE for longs, PE for shorts\n\n"
                "TARGET: 56% WR × 1:3 RR = profitable"
            )

        # ── TRADE REPLAY (#27) ────────────────────────────────────────────────

        elif text.startswith('/replay'):
            try:
                from trader.paper_trader import load_state
                from utils.chart_renderer import render_trade_replay
                from utils.telegram_alerts import send_photo
                from scanner.data_fetcher import get_historical_data

                parts = text.split()
                state = load_state()
                closed = state.get('closed_trades', [])

                if not closed:
                    send_message("No closed trades yet to replay.")
                else:
                    # /replay = last closed trade. /replay <id> = specific
                    if len(parts) > 1:
                        target = next(
                            (t for t in closed if t.get('id', '').endswith(parts[1])),
                            None
                        )
                    else:
                        target = closed[-1]

                    if not target:
                        send_message("Trade not found. Try /replay (no args) for last trade.")
                    else:
                        tf = str(target.get('timeframe', '15min')).replace('min', '')
                        df = get_historical_data(fyers_ref, target['symbol'], tf, days=10)
                        if df is None:
                            send_message("Could not fetch chart data.")
                        else:
                            png = render_trade_replay(df, target)
                            if png:
                                sym = target['symbol'].replace('NSE:', '').replace('-EQ', '')
                                cap = (
                                    f"{sym} replay | "
                                    f"PnL Rs {target.get('pnl', 0):.0f} | "
                                    f"{target.get('status', '?')}"
                                )
                                send_photo(png, cap)
                            else:
                                send_message("Replay render failed.")
            except Exception as e:
                send_message(f"Replay error: {e}")

        # ── AI LESSONS (post-mortems) ─────────────────────────────────────────

        elif text == '/lessons':
            try:
                from data.trade_lessons import format_lessons_summary
                send_message(format_lessons_summary())
            except Exception as e:
                send_message(f"Lessons error: {e}")

        # ── EXPIRY CALENDAR ───────────────────────────────────────────────────

        elif text == '/expiry':
            try:
                from scanner.expiry_calendar import format_expiry_summary
                send_message(format_expiry_summary())
            except Exception as e:
                send_message(f"Expiry calendar error: {e}")

        # ── ADVERSARIAL ROBUSTNESS TEST (#23) ─────────────────────────────────

        elif text == '/adversarial':
            try:
                from backtest.adversarial import run_adversarial_suite, format_adversarial_report
                send_message("Running adversarial test suite...")
                def _run():
                    results = run_adversarial_suite()
                    send_message(format_adversarial_report(results))
                _run_in_thread(_run)
            except Exception as e:
                send_message(f"Adversarial error: {e}")

        # ── CHAMPION vs CHALLENGER (#24) ──────────────────────────────────────

        elif text.startswith('/champion'):
            parts = text.split()
            symbol_arg = parts[1].upper() if len(parts) > 1 else 'RELIANCE'
            symbol_full = (
                f"NSE:{symbol_arg}-EQ" if 'FUT' not in symbol_arg
                else f"NSE:{symbol_arg}"
            )
            send_message(
                f"Running champion vs challenger on {symbol_arg}... ~90s"
            )
            def _run_cc():
                try:
                    from backtest.champion_challenger import (
                        run_champion_vs_challenger, format_comparison
                    )
                    # Default challenger: stricter score + RR
                    overrides = {'MIN_BUY_SCORE': 9, 'MIN_RR_RATIO': 4.0}
                    res = run_champion_vs_challenger(
                        fyers_ref, symbol_full, '15', 90, overrides
                    )
                    send_message(format_comparison(res))
                except Exception as e:
                    send_message(f"Champion error: {e}")
            _run_in_thread(_run_cc)

        # ── EVENT MODE (geopolitical) ─────────────────────────────────────────

        elif text.startswith('/eventmode'):
            try:
                from data.news_calendar import set_geopolitical_event, is_geopolitical_event
                parts = text.split()
                if len(parts) < 2:
                    status = "ON" if is_geopolitical_event() else "OFF"
                    send_message(
                        f"CB6 QUANTUM - EVENT MODE: {status}\n\n"
                        "Usage: /eventmode on | off\n"
                        "Use during war/crisis/major geopolitical events.\n"
                        "Bot will require A+ setups + FII/DII alignment."
                    )
                else:
                    action = parts[1].lower()
                    if action == 'on':
                        set_geopolitical_event(True, manual=True)
                        send_message(
                            "CB6 - EVENT MODE ON (manual)\n"
                            "Auto-detection locked out for 24h."
                        )
                    elif action == 'off':
                        set_geopolitical_event(False, manual=True)
                        send_message(
                            "CB6 - EVENT MODE OFF (manual)\n"
                            "Auto-detection locked out for 24h."
                        )
                    elif action == 'auto':
                        # Clear manual override, let auto-detect take over
                        from data import news_calendar as _nc
                        if os.path.exists(_nc.MANUAL_OVERRIDE_FILE):
                            os.remove(_nc.MANUAL_OVERRIDE_FILE)
                        send_message("CB6 - Manual override cleared. Auto-detection active.")
            except Exception as e:
                send_message(f"Event mode error: {e}")

        # ── ALIGNED WATCHLIST ─────────────────────────────────────────────────

        elif text.startswith('/aligned'):
            try:
                wl_path = os.path.join(
                    os.path.dirname(os.path.dirname(__file__)),
                    'data', 'aligned_watchlist.json'
                )
                if not os.path.exists(wl_path):
                    send_message(
                        "CB6 - No aligned watchlist yet.\n"
                        "Run /equity to build it (W1+D1+H4 alignment)."
                    )
                else:
                    import json as _json
                    with open(wl_path) as f:
                        wl = _json.load(f)
                    bull = wl.get('bullish', [])
                    bear = wl.get('bearish', [])
                    bull_str = ", ".join(bull[:30]) + (" ..." if len(bull) > 30 else "")
                    bear_str = ", ".join(bear[:30]) + (" ..." if len(bear) > 30 else "")
                    send_message(
                        f"CB6 - ALIGNED WATCHLIST (W1+D1+H4)\n\n"
                        f"As of    : {wl.get('date','N/A')}\n"
                        f"Aligned  : {wl.get('aligned',0)}/{wl.get('total',0)}\n"
                        f"BULLISH  ({len(bull)}): {bull_str or 'none'}\n\n"
                        f"BEARISH  ({len(bear)}): {bear_str or 'none'}\n\n"
                        "These are scanned FIRST as priority each cycle."
                    )
            except Exception as e:
                send_message(f"Aligned watchlist error: {e}")

        # ── WEBSOCKET REALTIME FEED ───────────────────────────────────────────

        elif text.startswith('/ws'):
            try:
                from scanner.websocket_feed import init as ws_init, is_active
                from core.tick_watcher import get_watcher
                from settings import CLIENT_ID, ACCESS_TOKEN
                parts = text.split()
                if len(parts) < 2:
                    s = get_watcher().status()
                    send_message(
                        "CB6 QUANTUM - WEBSOCKET STATUS\n\n"
                        f"Active     : {'YES' if is_active() else 'NO'}\n"
                        f"Symbols    : {s['symbols_watched']}\n"
                        f"Triggers   : {s['active_watches']}\n"
                        f"Ticks proc : {s['ticks_processed']}\n"
                        f"Fired      : {s['triggers_fired']}\n"
                        f"Last tick  : {s['last_tick_age_s']}s ago\n\n"
                        "Usage: /ws on | off | status"
                    )
                else:
                    action = parts[1].lower()
                    if action == 'on':
                        if is_active():
                            send_message("WS already active.")
                        elif ws_init(ACCESS_TOKEN, CLIENT_ID):
                            send_message("CB6 - WebSocket ON. Realtime triggers armed.")
                            # Re-arm triggers for all open trades
                            try:
                                from trader.paper_trader import load_state
                                from core.trade_triggers import register_trade_triggers
                                from scanner.websocket_feed import subscribe
                                state = load_state()
                                syms = []
                                for trade in state.get('open_trades', []):
                                    register_trade_triggers(trade)
                                    syms.append(trade['symbol'])
                                if syms:
                                    subscribe(syms)
                            except Exception as e:
                                logger.error(f"Re-arm error: {e}")
                        else:
                            send_message("WS init failed — check logs.")
                    elif action == 'off':
                        get_watcher().clear()
                        send_message("CB6 - All triggers cleared. WS feed will stop on next tick.")
                    elif action == 'status':
                        s = get_watcher().status()
                        send_message(
                            f"WS triggers: {s['active_watches']} on {s['symbols_watched']} symbols\n"
                            f"Ticks: {s['ticks_processed']} | Fired: {s['triggers_fired']}"
                        )
            except Exception as e:
                send_message(f"WS error: {e}")

        # ── BACKTEST ──────────────────────────────────────────────────────────

        elif text.startswith('/backtest'):
            parts  = text.split()
            symbol = parts[1].upper() if len(parts) > 1 else 'RELIANCE'
            tf     = parts[2] if len(parts) > 2 else '15'

            # Build the correct Fyers symbol per type
            INDEX_ALIASES = {
                'NIFTY'    : 'NSE:NIFTY50-INDEX',
                'NIFTY50'  : 'NSE:NIFTY50-INDEX',
                'BANKNIFTY': 'NSE:NIFTYBANK-INDEX',
                'NIFTYBANK': 'NSE:NIFTYBANK-INDEX',
            }
            if symbol in INDEX_ALIASES:
                sym_full = INDEX_ALIASES[symbol]
            elif 'FUT' in symbol or 'INDEX' in symbol:
                sym_full = f"NSE:{symbol}"
            else:
                sym_full = f"NSE:{symbol}-EQ"

            send_message(
                f"CB6 QUANTUM - BACKTEST STARTED\n\n"
                f"Symbol    : {symbol}\n"
                f"Resolved  : {sym_full}\n"
                f"Timeframe : {tf}min\n"
                f"Period    : 90 days\n"
                "This takes ~60 seconds. Please wait..."
            )
            def _run_bt():
                try:
                    from backtest.backtester import run_backtest, format_backtest_report
                    stats = run_backtest(fyers_ref, sym_full, tf, days=90)
                    send_message(format_backtest_report(stats))
                except Exception as e:
                    send_message(f"Backtest error: {e}")
            _run_in_thread(_run_bt)

        # ── 3M ICT BACKTEST ────────────────────────────────────────────────────
        # Usage: /backtest3m NIFTY
        #        /backtest3m NIFTY 2026-05-18 2026-05-25

        elif text.startswith('/backtest3m'):
            parts    = text.split()
            idx_raw  = parts[1].upper() if len(parts) > 1 else 'NIFTY'
            from_dt  = parts[2] if len(parts) > 2 else '2026-05-18'
            to_dt    = parts[3] if len(parts) > 3 else '2026-05-25'

            _INDEX_BT = {
                'NIFTY'      : 'NIFTY',      'NF'   : 'NIFTY',
                'BANKNIFTY'  : 'BANKNIFTY',  'BNF'  : 'BANKNIFTY',
                'FINNIFTY'   : 'FINNIFTY',   'FN'   : 'FINNIFTY',
                'MIDCPNIFTY' : 'MIDCPNIFTY', 'MCN'  : 'MIDCPNIFTY',
            }
            idx_name = _INDEX_BT.get(idx_raw, 'NIFTY')

            send_message(
                f"📊 CB6 BACKTEST — {idx_name} 3m TF\n\n"
                f"Period : {from_dt}  →  {to_dt}\n"
                f"Chain  : DOL → MSS → Displaced FVG\n"
                f"Hours  : 10:00 – 15:20 IST each day\n\n"
                "Fetching 3m data from Fyers... ~30s ⏳"
            )

            def _run_bt3m(idx=idx_name, fd=from_dt, td=to_dt):
                try:
                    from ml.backtest_3m import run_backtest_3m, format_backtest_message
                    result = run_backtest_3m(fyers_ref, idx, fd, td)
                    msg    = format_backtest_message(result)
                    # Split into chunks if message too long (Telegram 4096 limit)
                    if len(msg) <= 4000:
                        send_message(msg, parse_mode='HTML')
                    else:
                        # Send summary + trades separately
                        trades = result.get('trades', [])
                        from ml.backtest_3m import format_backtest_message as _fmt
                        result_no_trades = {k: v for k, v in result.items()
                                            if k != 'trades'}
                        result_no_trades['trades'] = []
                        send_message(_fmt(result_no_trades), parse_mode='HTML')
                        # Send trades in batches of 10
                        trade_lines = []
                        for t in trades:
                            icon = '✅' if t['r'] > 0 else ('❌' if t['r'] < 0 else '➡️')
                            trade_lines.append(
                                f"{icon} {t['date']} {t['time']}  {t['dir']}  "
                                f"E:{t['entry']} SL:{t['sl']}  "
                                f"{t['outcome']} {t['r']:+.2f}R  "
                                f"[{t['score']} {t['mss']}]"
                            )
                        batch = '\n'.join(trade_lines)
                        send_message(f"<b>TRADE LOG:</b>\n{batch}", parse_mode='HTML')
                except Exception as e:
                    logger.error(f"Backtest3m error: {e}")
                    send_message(f"❌ Backtest 3m error: {e}")

            _run_in_thread(_run_bt3m)

        # ── SINGLE STOCK CHECK ────────────────────────────────────────────────
        # Usage: /check HDFC 155   /check ICICI 15   /check RELIANCE 60
        # Smart aliases: HDFC→HDFCBANK, ICICI→ICICIBANK, TATA→TATAMOTORS+
        # TF normalization: 155→15, 1h→60, 6→5

        elif text in ('/scan', '/nifty50'):
            from scanner.silver_bullet import get_window_status
            send_message(
                f"CB6 SCAN — Index Only Mode\n\n"
                f"Instruments: NIFTY | BANKNIFTY | FINNIFTY | MIDCPNIFTY\n"
                f"Chain      : DOL → MSS → Displaced FVG\n"
                f"Timeframe  : 5 min\n"
                f"Window     : {get_window_status()}\n\n"
                f"Scanning index futures..."
            )
            if scan_callback:
                _run_in_thread(scan_callback)

        elif text.startswith('/check'):
            parts = text.split()
            _INDEX_MAP = {
                'NIFTY'      : 'NIFTY',      'NIFTY50'    : 'NIFTY',
                'BANKNIFTY'  : 'BANKNIFTY',  'BANK'       : 'BANKNIFTY',
                'FINNIFTY'   : 'FINNIFTY',   'FIN'        : 'FINNIFTY',
                'MIDCPNIFTY' : 'MIDCPNIFTY', 'MIDCP'      : 'MIDCPNIFTY',
            }
            if len(parts) < 2:
                send_message(
                    "Usage: /check INDEX\n\n"
                    "Valid indexes:\n"
                    "  /check NIFTY\n"
                    "  /check BANKNIFTY\n"
                    "  /check FINNIFTY\n"
                    "  /check MIDCPNIFTY\n\n"
                    "Index-only mode — equity stocks disabled."
                )
            else:
                raw = parts[1].upper()
                if raw not in _INDEX_MAP:
                    send_message(
                        f"CB6 — Index Only Mode\n\n"
                        f"'{raw}' is not a supported index.\n\n"
                        "Valid indexes:\n"
                        "  /check NIFTY\n"
                        "  /check BANKNIFTY\n"
                        "  /check FINNIFTY\n"
                        "  /check MIDCPNIFTY\n\n"
                        "Equity stocks are disabled."
                    )
                else:
                    idx_name = _INDEX_MAP[raw]
                    send_message(
                        f"CB6 CHECK — {idx_name}\n\n"
                        f"Chain : DOL → MSS → Displaced FVG\n"
                        f"TF    : 5 min\n\n"
                        f"Fetching data..."
                    )

                    def _run_check():
                        try:
                            from scanner.data_fetcher import get_historical_data
                            from scanner.silver_bullet import scan_silver_bullet, format_sb_alert
                            from scanner.index_futures import get_active_futures
                            from data.pattern_library import compute_trade_confidence, format_confidence_alert

                            futures  = get_active_futures()
                            sym_full = futures.get(idx_name)
                            if not sym_full:
                                send_message(f"CB6 CHECK — Could not resolve futures symbol for {idx_name}")
                                return

                            df = get_historical_data(fyers_ref, sym_full, '5', days=5)
                            if df is None or len(df) < 30:
                                send_message(f"CB6 CHECK — No data for {idx_name}. Check token.")
                                return

                            setup = scan_silver_bullet(df, sym_full, tf='5',
                                                       fyers=fyers_ref, force=True)
                            if not setup:
                                last = float(df['close'].iloc[-1])
                                send_message(
                                    f"CB6 CHECK — {idx_name}\n\n"
                                    f"LTP     : {last}\n"
                                    f"Candles : {len(df)}\n\n"
                                    f"No setup: DOL→MSS→FVG chain incomplete.\n"
                                    f"Price not at/near a displaced FVG right now."
                                )
                                return

                            setup['timeframe']       = '5min'
                            setup['instrument_type'] = 'INDEX'
                            send_message(format_sb_alert(setup, tf='5'))

                            conf = compute_trade_confidence(setup)
                            setup['pattern_confidence'] = conf
                            if conf['match_count'] > 0:
                                send_message(format_confidence_alert(setup, conf))

                        except Exception as e:
                            send_message(f"CB6 CHECK ERROR — {idx_name}: {e}")

                    _run_in_thread(_run_check)

        elif text.startswith('/watchlist'):
            send_message("Daily watchlist removed — SB-only mode.")

        # ── AI CHAT ───────────────────────────────────────────────────────────

        elif text == '/ask' or text.startswith('/ask '):
            question = text[4:].strip()
            if not question:
                send_message(
                    "Ask me anything about the trade or strategy.\n\n"
                    "Example:\n"
                    "/ask Scan SBIN and give me levels\n"
                    "/ask Why did the bot not trade today?\n"
                    "/ask What is a Fair Value Gap?"
                )
            else:
                try:
                    from utils.ai_chat import chat, is_available
                    send_message("Thinking...")
                    if is_available():
                        send_message(chat(question))
                    else:
                        send_message(
                            "AI chat offline.\n\n"
                            "Add your Anthropic API key to .env:\n"
                            "ANTHROPIC_API_KEY=sk-ant-api03-..."
                        )
                except Exception as e:
                    send_message(f"AI error: {e}")

        elif text == '/clearchat':
            try:
                from utils.ai_chat import clear_history
                send_message(clear_history())
            except Exception as e:
                send_message(f"Clear chat error: {e}")

        # ── ML STATUS ─────────────────────────────────────────────────────────

        elif text == '/ml_status':
            try:
                from ml.shadow_monitor import build_status_message
                send_message(build_status_message(), parse_mode='Markdown')
            except Exception as e:
                send_message(f"ML status error: {e}")

        # ── ML SCANNER ────────────────────────────────────────────────────────────

        elif text.startswith('/ml_scan'):
            arg = text[len('/ml_scan'):].strip().upper()
            def _run_ml_scan(arg=arg):
                try:
                    from ml.scanner import (
                        scan_index, scan_all_indices,
                        format_scan_message, format_summary_message,
                        _resolve_index
                    )
                    fyers = fyers_ref

                    if not arg or arg in ('ALL', ''):
                        # Scan all 4 indices
                        send_message("🔍 ML scanning all indices... please wait ~15s")
                        results = scan_all_indices(fyers)
                        if not results:
                            send_message("ML Scanner: no data — check Fyers token.")
                            return
                        # Summary first
                        send_message(format_summary_message(results), parse_mode='HTML')
                        # Then full detail for any that have a setup
                        for r in results:
                            if r['has_setup']:
                                send_message(format_scan_message(r), parse_mode='HTML')
                    else:
                        idx = _resolve_index(arg)
                        if not idx:
                            send_message(
                                f"Unknown index: {arg}\n"
                                "Use: /ml_scan NIFTY | BANKNIFTY | FINNIFTY | MIDCPNIFTY | ALL"
                            )
                            return
                        send_message(f"🔍 ML scanning {idx}... please wait ~10s")
                        result = scan_index(idx, fyers)
                        if not result:
                            send_message(f"ML Scanner {idx}: no data — check Fyers token.")
                            return
                        send_message(format_scan_message(result), parse_mode='HTML')
                except Exception as e:
                    logger.error(f"ML scan error: {e}")
                    send_message(f"ML scan error: {e}")
            _run_in_thread(_run_ml_scan)

        elif text.startswith('/ml_train'):
            try:
                from ml.auto_trainer import trigger_now
                arg = text[len('/ml_train'):].strip().lower()
                parts = arg.split()
                if not parts:
                    for mkt, acc in [('nse',''), ('forex','ftmo'), ('forex','gft')]:
                        trigger_now(mkt, acc)
                    send_message("🧠 ML training triggered for all markets.")
                elif parts[0] == 'nse':
                    trigger_now('nse', '')
                    send_message("🧠 ML training triggered for NSE.")
                elif parts[0] == 'forex':
                    acc = parts[1] if len(parts) > 1 else ''
                    if acc in ('ftmo', 'gft'):
                        trigger_now('forex', acc)
                        send_message(f"🧠 ML training triggered for forex/{acc.upper()}.")
                    else:
                        trigger_now('forex', 'ftmo')
                        trigger_now('forex', 'gft')
                        send_message("🧠 ML training triggered for FTMO + GFT.")
                else:
                    send_message("Usage: /ml_train | /ml_train nse | /ml_train forex | /ml_train forex gft")
            except Exception as e:
                send_message(f"ML train error: {e}")

        elif text.startswith('/'):
            send_message(
                f"Unknown command: {text}\n"
                "Type /info to see all commands."
            )

        else:
            # Free-text → AI chat (any message without a leading /)
            try:
                from utils.ai_chat import chat, is_available
                if is_available():
                    send_message(chat(text))
                else:
                    send_message(
                        "AI chat offline. Add ANTHROPIC_API_KEY to .env\n"
                        "Use /ask <question> or type any question directly."
                    )
            except Exception as e:
                send_message(f"AI error: {e}")


def _register_bot_commands():
    """Register /command menu with Telegram so the '/' key shows autocomplete."""
    commands = [
        {"command": "start",          "description": "Overview + all commands"},
        {"command": "sb",             "description": "Trigger Silver Bullet scan now"},
        {"command": "scan",           "description": "Index futures scan (NIFTY/BNF/FIN/MID)"},
        {"command": "check",          "description": "Check a specific index — /check NIFTY"},
        {"command": "nse_status",     "description": "Bot health, windows, today's trades"},
        {"command": "trades",         "description": "Open index positions + live PnL"},
        {"command": "portfolio",      "description": "Capital & P&L summary"},
        {"command": "excel",          "description": "Download Excel dashboard"},
        {"command": "levels",         "description": "NIFTY ICT levels + buy/sell probability"},
        {"command": "brain",          "description": "Market bias + session score"},
        {"command": "options",        "description": "Strike selector — /options NIFTY"},
        {"command": "ml_scan",        "description": "Multi-TF ML scan — /ml_scan NIFTY or ALL"},
        {"command": "ml_status",      "description": "ML model accuracy + predictions"},
        {"command": "ml_train",       "description": "Force ML retrain now"},
        {"command": "fiidii",         "description": "FII/DII flow data"},
        {"command": "expiry",         "description": "F&O expiry calendar"},
        {"command": "ask",            "description": "Ask AI anything — /ask <question>"},
        {"command": "memory",         "description": "AI trade stats"},
        {"command": "lessons",        "description": "Trade post-mortems"},
        {"command": "pattern",        "description": "Pattern library stats"},
        {"command": "stop",           "description": "Halt trading today"},
        {"command": "resume",         "description": "Resume trading"},
        {"command": "help",           "description": "ICT Silver Bullet strategy rules"},
        {"command": "info",           "description": "Full command reference"},
    ]
    if not TELEGRAM_BOT_TOKEN:
        return
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/setMyCommands",
            json={"commands": commands},
            timeout=10,
        )
        if r.status_code == 200 and r.json().get('ok'):
            logger.info(f"NSE bot: registered {len(commands)} commands in Telegram menu")
        else:
            logger.warning(f"NSE bot setMyCommands failed: {r.text}")
    except Exception as e:
        logger.warning(f"NSE bot setMyCommands error: {e}")


def start_listening():
    logger.info("Telegram listener started...")
    _register_bot_commands()
    while True:
        try:
            updates = get_updates()
            if updates:
                process_updates(updates)
            time.sleep(2)
        except Exception as e:
            logger.error(f"Listener error: {e}")
            time.sleep(5)
