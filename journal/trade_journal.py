# journal/trade_journal.py â€” CB6 QUANTUM Excel Trade Journal
import os
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
sys.path.append(os.path.abspath(
    os.path.join(os.path.dirname(__file__), '..')
))
from utils.logger import logger

JOURNAL_PATH = os.path.join(
    os.path.dirname(__file__), '..', 'journal', 'cb6_trades.xlsx'
)

# Colors
GREEN  = "00B050"
RED    = "FF0000"
BLUE   = "0070C0"
YELLOW = "FFD700"
GREY   = "F2F2F2"
WHITE  = "FFFFFF"
DARK   = "1F1F1F"

def get_border():
    side = Side(style='thin', color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def create_journal():
    """Create Excel journal if not exists"""
    try:
        if os.path.exists(JOURNAL_PATH):
            wb = openpyxl.load_workbook(JOURNAL_PATH)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet
            wb.remove(wb.active)

        # Create sheets if not exist
        sheets = ['Trades', 'Daily', 'Weekly', 'Monthly', 'Summary']
        for sheet in sheets:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)

        # Setup Trades sheet headers
        ws = wb['Trades']
        if ws.max_row == 1 and ws['A1'].value is None:
            headers = [
                'Trade ID', 'Date', 'Time', 'Symbol',
                'Timeframe', 'Entry', 'Exit', 'SL',
                'T1', 'T2', 'T3', 'Quantity',
                'Position Value', 'PnL', 'Result',
                'Exit Reason', 'RR Ratio', 'Duration',
                'B1', 'B2', 'Neckline'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font      = Font(bold=True, color=WHITE)
                cell.fill      = PatternFill("solid", fgColor=BLUE)
                cell.alignment = Alignment(horizontal='center')
                cell.border    = get_border()
            
            # Set column widths
            widths = [15,12,10,15,10,10,10,10,10,10,10,10,15,10,10,12,10,10,10,10,10]
            for col, width in enumerate(widths, 1):
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col)
                ].width = width

        # Setup Summary sheet
        ws_sum = wb['Summary']
        if ws_sum['A1'].value is None:
            setup_summary_sheet(ws_sum)

        wb.save(JOURNAL_PATH)
        logger.info("Excel journal ready")
        return True

    except Exception as e:
        logger.error(f"Create journal error: {e}")
        return False


def setup_summary_sheet(ws):
    """Setup summary sheet with formulas"""
    title_font = Font(bold=True, size=14, color=WHITE)
    header_font = Font(bold=True, color=WHITE)

    # Title
    ws['A1'] = 'CB6 QUANTUM - TRADE SUMMARY'
    ws['A1'].font      = Font(bold=True, size=16, color=WHITE)
    ws['A1'].fill      = PatternFill("solid", fgColor=DARK)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:D1')

    # Stats labels
    stats = [
        ('A3', 'METRIC',          'VALUE'),
        ('A4', 'Total Trades',    "=COUNTA(Trades!A2:A10000)-1"),
        ('A5', 'Total Wins',      "=COUNTIF(Trades!O2:O10000,\"WIN\")"),
        ('A6', 'Total Losses',    "=COUNTIF(Trades!O2:O10000,\"LOSS\")"),
        ('A7', 'Win Rate %',      "=IF(B4>0,ROUND(B5/B4*100,1),0)"),
        ('A8', 'Total PnL',       "=SUM(Trades!N2:N10000)"),
        ('A9', 'Avg Profit',      "=IF(B5>0,AVERAGEIF(Trades!O2:O10000,\"WIN\",Trades!N2:N10000),0)"),
        ('A10','Avg Loss',        "=IF(B6>0,AVERAGEIF(Trades!O2:O10000,\"LOSS\",Trades!N2:N10000),0)"),
        ('A11','Best Trade',      "=MAX(Trades!N2:N10000)"),
        ('A12','Worst Trade',     "=MIN(Trades!N2:N10000)"),
        ('A13','Profit Factor',   "=IF(ABS(B10)>0,ABS(B9/B10),0)"),
    ]

    for label_cell, label, value in stats:
        col_a = ws[label_cell]
        col_a.value     = label
        col_a.font      = Font(bold=True)
        col_a.fill      = PatternFill("solid", fgColor=GREY)
        col_a.border    = get_border()

        val_cell = label_cell.replace('A', 'B')
        ws[val_cell] = value
        ws[val_cell].border = get_border()
        ws[val_cell].alignment = Alignment(horizontal='center')


def log_trade(trade):
    """Log a completed trade to Excel"""
    try:
        create_journal()
        wb = openpyxl.load_workbook(JOURNAL_PATH)
        ws = wb['Trades']

        # Find next empty row
        next_row = ws.max_row + 1
        if next_row == 2 and ws['A2'].value is None:
            next_row = 2

        # Calculate duration
        try:
            entry_dt = datetime.strptime(
                trade['entry_time'], '%Y-%m-%d %H:%M:%S'
            )
            exit_dt  = datetime.strptime(
                trade['exit_time'], '%Y-%m-%d %H:%M:%S'
            )
            duration = str(exit_dt - entry_dt)
        except:
            duration = "N/A"

        result = "WIN" if trade['pnl'] > 0 else "LOSS"

        # Row data
        row_data = [
            trade['id'],
            trade['entry_time'].split(' ')[0],
            trade['entry_time'].split(' ')[1],
            trade['symbol'].replace("NSE:","").replace("-EQ",""),
            trade.get('timeframe', '15min'),
            trade['entry_price'],
            trade.get('exit_price', 0),
            trade['stop_loss'],
            trade['target1'],
            trade['target2'],
            trade['target3'],
            trade['quantity'],
            trade['position_value'],
            round(trade['pnl'], 2),
            result,
            trade.get('status', 'N/A'),
            trade['rr_ratio'],
            duration,
            trade.get('b1_price', 0),
            trade.get('b2_price', 0),
            trade.get('neck_price', 0),
        ]

        # Write row
        for col, value in enumerate(row_data, 1):
            cell        = ws.cell(row=next_row, column=col, value=value)
            cell.border = get_border()
            cell.alignment = Alignment(horizontal='center')

            # Color PnL cell
            if col == 14:
                if trade['pnl'] > 0:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(color=GREEN, bold=True)
                else:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(color=RED, bold=True)

            # Color result cell
            if col == 15:
                if result == "WIN":
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(color=GREEN, bold=True)
                else:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(color=RED, bold=True)

            # Alternate row colors
            if next_row % 2 == 0 and col not in [14, 15]:
                cell.fill = PatternFill("solid", fgColor=GREY)

        wb.save(JOURNAL_PATH)
        logger.info(f"Trade logged to Excel: {trade['id']}")
        return True

    except Exception as e:
        logger.error(f"Log trade error: {e}")
        return False


def get_weekly_summary():
    """Get weekly performance summary"""
    try:
        if not os.path.exists(JOURNAL_PATH):
            return None

        wb = openpyxl.load_workbook(JOURNAL_PATH)
        ws = wb['Trades']

        trades = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                trades.append(row)

        if not trades:
            return None

        total  = len(trades)
        wins   = sum(1 for t in trades if t[14] == 'WIN')
        losses = sum(1 for t in trades if t[14] == 'LOSS')
        pnl    = sum(t[13] for t in trades if t[13])

        return {
            'total' : total,
            'wins'  : wins,
            'losses': losses,
            'pnl'   : pnl,
            'winrate': round(wins/total*100, 1) if total > 0 else 0
        }

    except Exception as e:
        logger.error(f"Weekly summary error: {e}")
        return None
